import matplotlib.pyplot as plotter
from openpyxl import load_workbook
import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter.colorchooser import askcolor
from statistics import mean, stdev 
from csv import reader as csv_reader
from os.path import split as path_split

def print_help():
    print("Plot generator from excel options (2023) Mateusz Ferenc:")
    print("\tq, quit, exit, leave - to exit")
    print("\tload - open file dialog window to select file to load data from")
    print("\tunload - delete loaded file from temporary memory")
    print("\tselect <row-start> <row-end> <column-start> <column-end> or select <row-start> <row-end> <column> - select data within given range (column accepts excel column representation) from excel file / select data within given row range from <column> from csv file\nWhen used command \"search\" were used and found accessible data, then cache will be used instead of directly reading data from loaded file")
    print("\tdata or data <data-set> - print loaded datasets (only informations) / print contents of <data-set>")
    print("\tclear or clear <data-set> or clear plot - clear whole database / clear <data-set> entry / clear plot properties")
    print("\tconfig plot or config data or config <data-set> - configure plotter properties / configure each dataset plot properties / configure <data-set> plot properties")
    print("\tsearch - search for accessible data within loaded file")
    print("\tgenerate - generate plot from given data")

def excel_column_to_num(col: str) -> (int | None):
    if type(col) is not str:
        return None
    col_num = 0
    for c in range(len(col)):
        ascii_ = ord(col[-(1 + c)])
        if 65 <= ascii_ <= 90:
            col_num += (ascii_ - 64) * pow(26, c)   # value in range 1 - 26
        else:
            return None
    return col_num

def num_to_excel_column(col: int) -> (str | None):
    if type(col) is not int:
        return None
    x = (col % 26)
    r = col // 26
    if (r == 0):
        return chr(65 + x)      # value in range 0 - 25
    return num_to_excel_column(r - 1) + chr(65 + x)

def input_catch(text: str, do_return = None) -> (str | None):
    try:
        temp = input(text)
        if len(temp):
            return temp
        else:
            raise Exception("Empty Input")
    except:
        return do_return

def is_float(element) -> bool:
    if element is None: 
        return False
    try:
        float(element)
        return True
    except ValueError:
        return False

def search_for_data(file_handler, file_type: str) -> (dict | None):
    if file_type in supported_file_types['excel'] and file_handler is not None:
        max_row, max_col = current_sheet.max_row, current_sheet.max_column
        temp = {}
        for col in range(1, max_col + 1):
            start, stop = None, None
            temp_data = []
            for row in range(1, max_row + 1):
                if row % 100 == 0:
                    print(end='\x1b[2K')
                    print("Please wait" + "."*((row//100)%10), end="\r")
                value = current_sheet.cell(row=row, column=col).value
                if value is not None and is_float(value):
                    if start is None:
                        start = row
                    temp_data.append(round(float(value), 4))
                else:
                    if start is not None:
                        stop = row - 1
                if start is not None and stop is not None:
                    temp[f"{start}:{stop}:{num_to_excel_column(col - 1)}:{num_to_excel_column(col - 1)}"] = temp_data
                    temp_data = []
                    start, stop = None, None
        print(end='\x1b[2K')
        return temp if temp != {} else None
    elif file_type in supported_file_types['csv'] and file_handler is not None:
        ncol = len(next(workbook))
        csv_file.seek(0)
        temp = {}
        for col in range(ncol):
            start, stop = None, None
            temp_data = []
            for row_num, row in enumerate(workbook):
                if row_num % 100 == 0:
                    print(end='\x1b[2K')
                    print("Please wait" + "."*((row_num//100)%10), end="\r")
                value = row[col]
                if value is not None and is_float(value):
                    if start is None:
                        start = row_num
                    temp_data.append(round(float(value), 4))
                else:
                    if start is not None:
                        stop = row_num - 1
                if start is not None and stop is not None:
                    temp[f"{start}:{stop}:{num_to_excel_column(col)}"] = temp_data
                    temp_data = []
                    start, stop = None, None
            if start is not None and stop is None:
                temp[f"{start}:{row_num}:{num_to_excel_column(col)}"] = temp_data
        print(end='\x1b[2K')
        return temp if temp != {} else None
    else:
        return None

if __name__ == "__main__":
    root = tk.Tk()
    root.attributes('-topmost', True)
    root.withdraw()

    workbook = None
    current_sheet = None
    csv_file = None
    data = {}
    plot_properties = {
        "dpi": 500,
        "xlabel": None,
        "ylabel": None,
        "label_weight": "light",
        "label_style": "italic",
        "title": None,
        "title_weight": "bold",
        "title_style": "normal",
        "grid": False,
        "width": 18,
        "height": 10
    }
    plot_properties_ok = False
    current_file_type = ""
    current_filename = ""
    supported_file_types = {
        "excel": (".xlsx", ".xlsm", ".xltx", ".xltm"),
        "csv": ".csv"
    }
    fs_names = {
        "excel": "Excel files",
        "csv": "Comma Separated Values"
    }
    accessible_data = {}

    while(True):
        command = input_catch(">", "help")
        if command in ("h", "help"):
            print_help()
            continue
        if command.startswith("exec"): # debug-only
            try:
                exec(command.replace("exec ", ""))
            except Exception as e:
                print(e)
            continue
        command, command_data = command.split()[0].lower(), command.split()[1:]
        if command in ("q", "quit", "exit", "leave"):
            if workbook is not None:
                if current_file_type in supported_file_types['excel']:
                    workbook.close()
                elif current_file_type in supported_file_types['csv']:
                    csv_file.close()
            exit()
        elif command == "load":
            if len(command_data) == 0:
                ans = ""
                accessible_data = {}
                if workbook is not None:
                    ans = input_catch(f"{current_sheet.title} is currently loaded, unload? (y/n)", "n")
                if ans.lower() in ("y", "yes", "yep") or workbook is None:
                    path = askopenfilename(title="Choose excel file", filetypes=[(fs_names['excel'], " ".join(supported_file_types['excel'])), (fs_names['csv'], " ".join(supported_file_types['csv']))], parent=root)
                    try:
                        if len(path) == 0:
                            raise Exception("Empty path, no file selected.")
                        current_file_type = "." + path_split(path)[1].rsplit(".", 1)[1]
                        current_filename = path_split(path)[1].rsplit(".", 1)[0]
                        if current_file_type in supported_file_types["csv"]:
                            csv_file = open(path, 'r')
                            workbook = csv_reader(csv_file)
                        elif current_file_type in supported_file_types["excel"]:
                            workbook = load_workbook(path)
                            current_sheet = workbook.active
                        else:
                            raise Exception(f"{current_file_type} is unsupported..")
                    except FileNotFoundError:
                        print(f"{path} file not found...")
                    except Exception as e:
                        print(e)
                    else:
                        print(f"{current_filename}{current_file_type} loaded")
            else:
                print_help()
        elif command == "unload":
            if len(command_data) == 0:
                if workbook is not None:
                    if current_file_type in supported_file_types['excel']:
                        workbook.close()
                        workbook = None
                    elif current_file_type in supported_file_types['csv']:
                        workbook = None
                        csv_file.close()
                    print(f"{current_filename}{current_file_type} unloaded")
                    current_sheet = None
                    current_file_type = ""
                    current_filename = ""
                    accessible_data = {}
            else:
                print_help()
        elif command == "select":
            if workbook is None:
                print("First load file using \"load\" command")
            else:
                data_is_valid = False
                temp_data = []
                r0, r1, c0, c1 = None, None, None, None
                if len(command_data) == 3 and command_data[0].isdigit() and command_data[1].isdigit() and command_data[2].isalpha() and current_file_type in supported_file_types['csv']:
                    if accessible_data is not None:
                        for key in accessible_data.keys():
                            if f"{command_data[0]}:{command_data[1]}:{command_data[2]}" == key:
                                temp_data = accessible_data[key]
                                print(f"Cache used for selected range {key}")
                                data_is_valid = True
                                break
                    if data_is_valid == False:
                        ncol = len(next(workbook))
                        csv_file.seek(0)
                        col = excel_column_to_num(command_data[2])
                        if col is None or col > ncol:
                            print(f"{command_data[2]} is wrong column")
                            continue
                        try:
                            for row_num, row in enumerate(workbook):
                                if row_num > int(command_data[1]):
                                    break
                                if row_num < int(command_data[0]):
                                    continue
                                value = row[col]
                                if value is not None and is_float(value):
                                    temp_data.append(round(float(value), 4))
                                else:
                                    raise Exception(f"value: {value} in row: {row_num}, col: {col} is not int or float type\nType: {type(value)}\nRemoving selected data...")
                        except FileExistsError:
                            pass
                        else:
                            data_is_valid = True
                    c0 = c1 = command_data[2]
                    r0 = command_data[0]
                    r1 = command_data[1]
                elif len(command_data) == 4 and command_data[0].isdigit() and command_data[1].isdigit() and command_data[2].isalpha() and command_data[2].isalpha() and current_file_type in supported_file_types['excel']:
                    if accessible_data is not None:
                        for key in accessible_data.keys():
                            if f"{command_data[0]}:{command_data[1]}:{command_data[2]}:{command_data[3]}" == key:
                                temp_data = accessible_data[key]
                                print(f"Cache used for selected range {key}")
                                data_is_valid = True
                                break
                    if data_is_valid == False:
                        row_from, row_to, col_from, col_to = int(command_data[0]), int(command_data[1]), command_data[2], command_data[3]
                        ok = True
                        temp = excel_column_to_num(col_from)
                        col_from = temp
                        if temp is None:
                            ok = False
                        temp = excel_column_to_num(col_to)
                        col_to = temp
                        if temp is None:
                            ok =  False
                        if ok and (row_from <= row_to and col_from <= col_to):
                            try:
                                for row in range(row_from, row_to + 1):
                                    for col in range(col_from, col_to + 1):
                                        value = current_sheet.cell(row=row, column=col).value
                                        if is_float(value):
                                            temp_data.append(round(value, 4))
                                        else:
                                            raise Exception(f"value: {value} in row: {row}, col: {col} is not int or float type\nType: {type(value)}\nRemoving selected data...")
                            except Exception as e:
                                print(e)
                            else:
                                data_is_valid = True
                        else:
                            print("Selected range is not valid...\nTry again")
                    r0, r1, c0, c1 = command_data[0], command_data[1], command_data[2], command_data[3]
                else:
                    print("Use: \"select <row-start> <row-end> <column-start> <column-end>\" or \"select <row-start> <row-end> <column>\"")

                if data_is_valid:
                    datasets = [d for d in data.keys()]
                    print(f"Data is valid.\nLoaded datasets: {datasets}")
                    can_exit = False
                    while not can_exit:
                        data_name = input_catch("Enter name of new data set\n?")
                        if data_name not in ('plot', 'data') and data_name is not None:
                            data[data_name] = {}
                            can_exit = True
                        else:
                            print(f"{data_name} is protected name, use another..")
                    data[data_name]["data"] = temp_data
                    data[data_name]["info"] = {}
                    data[data_name]["info"]["sheet_name"] = current_filename
                    data[data_name]["info"]["row_from_to"] = f"{r0} : {r1}"
                    data[data_name]["info"]["col_from_to"] = f"{c0} : {c1}"
                    label = input_catch("Enter data label\n?")
                    data[data_name]["properties"] = {}
                    data[data_name]["properties"]["label"] = label
                    print("Color pick window summoned..")
                    color = askcolor(title="Choose plot color", parent=root)
                    data[data_name]["properties"]["color"] = color[1] if color != (None, None) else "#000000"
        elif command == "data":
            if len(command_data) == 0:
                if len(data):
                    for data_set in data.keys():
                        print(f"Dataset: \"{data_set}\" of size: {len(data[data_set]['data'])} elements")
                        print(f"Sheet name: {data[data_set]['info']['sheet_name']}, data selected= rows range {data[data_set]['info']['row_from_to']}, columns range {data[data_set]['info']['col_from_to']}")
                        print(f"Properties - label: {data[data_set]['properties']['label']}, color: {data[data_set]['properties']['color']}")
                        temp = data[data_set]['data']
                        print(f"Statistics - min: {round(min(temp), 4)}, max: {round(max(temp), 4)}, avg: {round(mean(temp), 4)}, stdev: {round(stdev(temp), 4)}")
                else:
                    print(f"No data loaded.\nUse \"select\" command to select new data")
            else:
                print_help()
        elif command == "clear":
            if len(command_data) == 0:
                data = {}
                plot_properties_ok = False
                accessible_data = {}
                print("All data cleared")
            elif len(command_data) == 1:
                if command_data[0] in data.keys():
                    del data[command_data[0]]
                    print(f"\"{command_data[0]}\" entries was removed from database")
                elif command_data[0] == "plot":
                    plot_properties_ok = False
                    print("Plot properties cleared.")
            else:
                print("Use: \"clear\" or \"clear <data-set>\" or \"clear plot\"")
        elif command == "config":
            if len(command_data) == 1:
                if command_data[0] == "plot":
                    can_exit = False
                    while not can_exit:
                        old = plot_properties["dpi"]
                        sel = input_catch(f"Enter plot DPI, or to preserve old \"{old}\" leave empty and hit enter\n?", old)
                        plot_properties["dpi"] = sel

                        old = plot_properties["xlabel"]
                        sel = input_catch(f"Enter x (horizontal) label name, or to preserve old \"{old}\" leave empty and hit enter\n?", old)
                        plot_properties["xlabel"] = sel

                        old = plot_properties["ylabel"]
                        sel = input_catch(f"Enter y (vertical) label name, or to preserve old \"{old}\" leave empty and hit enter\n?", old)
                        plot_properties["ylabel"] = sel

                        font_weights = ('ultralight', 'light', 'normal', 'regular', 'book', 'medium', 'roman', 'semibold', 'demibold', 'demi', 'bold', 'heavy', 'extra bold', 'black')

                        old = plot_properties["label_weight"]
                        sel = input_catch(f"Enter labels font weight (valid: {font_weights}), or to preserve old \"{old}\" leave empty and hit enter\n?", old)
                        if sel in font_weights:
                            plot_properties["label_weight"] = sel
                        else:
                            print(f"fontweight: {sel} is not valid..\n\"bold\" used as default")
                            plot_properties["label_weight"] = "bold"

                        font_styles = ('normal', 'italic', 'oblique')

                        old = plot_properties["label_style"]
                        sel = input_catch(f"Enter labels font style (valid: {font_styles}), or to preserve old \"{old}\" leave empty and hit enter\n?", old)
                        if sel in font_styles:
                            plot_properties["label_style"] = sel
                        else:
                            print(f"fontstyle: {sel} is not valid..\n\"light\" used as default")
                            plot_properties["label_style"] = "light"

                        old = plot_properties["title"]
                        sel = input_catch(f"Enter plot title, or to preserve old \"{old}\" leave empty and hit enter\n?", old)
                        plot_properties["title"] = sel

                        old = plot_properties["title_weight"]
                        sel = input_catch(f"Enter title font weight (valid: {font_weights}), or to preserve old \"{old}\" leave empty and hit enter\n?", old)
                        if sel in font_weights:
                            plot_properties["title_weight"] = sel
                        else:
                            print(f"fontweight: {sel} is not valid..\n\"bold\" used as default")
                            plot_properties["title_weight"] = "bold"

                        old = plot_properties["title_style"]
                        sel = input_catch(f"Enter labels font style (valid: {font_styles}), or to  preserve old \"{old}\" leave empty and hit enter\n?", old)
                        if sel in font_styles:
                            plot_properties["title_style"] = sel
                        else:
                            print(f"fontstyle: {sel} is not valid..\n\"normal\" used as default")
                            plot_properties["title_style"] = "normal"

                        old = plot_properties["grid"]
                        sel = input_catch(f"Enable grid? (valid: True, False), or to preserve old \"{old}\" leave empty and hit enter\n?", old)
                        if sel in ("True", "False"):
                            plot_properties["grid"] = sel
                        else:
                            print(f"option: {sel} is not valid..\n\"False\" used as default")

                        old = plot_properties['width']
                        sel = input_catch(f"Enter plot width (in cm), or to preserve old \"{old}\" leave empty and hit enter\n?", old)
                        plot_properties["width"] = sel
                        
                        old = plot_properties['height']
                        sel = input_catch(f"Enter plot height (in cm), or to preserve old \"{old}\" leave empty and hit enter\n?", old)
                        plot_properties["height"] = sel

                        print()
                        for key in plot_properties.keys():
                            print(f"{key}: {plot_properties[key]}")
                        ans = input_catch("Are you ok with these values? (y/n)?", "n")
                        if ans.lower() in ("y", "yes", "yep"):
                            can_exit = True
                            plot_properties_ok = True
                            print("OK!")
                elif len(data) == 0:
                    print("First you need to add dataset, using \"select\" command")
                elif command_data[0] == "data":
                    can_exit = False
                    while not can_exit:
                        for ds_name in data.keys():
                            print(f"{ds_name}:")
                            old = data[ds_name[0]]["properties"]["label"]
                            sel = input_catch(f"Enter data label for \"{ds_name}\" or to preserve old {old} leave empty and hit enter\n?", old)
                            data[ds_name]["properties"]["label"] = sel
                            old = data[ds_name[0]]["properties"]["color"]
                            selected_color = askcolor(title=f"Choose plot color, currently: {color}", parent=root)
                            data[ds_name]["properties"]["color"] = selected_color[1] if selected_color != (None, None) else old
                        for key in data.keys():
                            print(f"{key}:")
                            for k in data[key]["properties"].keys():
                                print(f"\t{k}: {data[key]['properties'][k]}")
                        ans = input_catch("Are you ok with these values? (y/n)?", "n")
                        if ans.lower() in ("y", "yes", "yep"):
                            can_exit = True
                            print("OK!")
                elif command_data[0] in data.keys():
                    can_exit = False
                    while not can_exit:
                        print(f"{command_data[0]}:")
                        old = data[command_data[0]]["properties"]["label"]
                        sel = input_catch(f"Enter data label for \"{ds_name}\" or to preserve old {old} leave empty and hit enter\n?", old)
                        data[data_name]["properties"]["label"] = sel
                        old = data[command_data[0]]["properties"]["color"]
                        selected_color = askcolor(title=f"Choose plot color, currently: {color}", parent=root)
                        data[data_name]["properties"]["color"] = selected_color[1] if selected_color != (None, None) else old
            else:
                print("Use: \"config plot\" or \"config data\" or \"config <data-set>\"")
        elif command == "search":
            if len(command_data) == 0:
                if len(accessible_data) == 0:
                    temp = search_for_data(file_handler=workbook, file_type=current_file_type)
                    if temp is None:
                        if current_file_type == "":
                            print("No file loaded, use \"load\" command first.")
                        else:
                            print(f"No valid data found in {current_filename} file..")
                        continue
                    else:
                        accessible_data = temp
                print(f"Found accessible data:")
                for r in accessible_data.keys():
                    print(f"{r} of size {int(r.split(':')[1]) - int(r.split(':')[0]) + 1} elements.")
            else:
                print_help()
        elif command == "generate":
            if plot_properties_ok and len(data):
                for data_set in data.keys():
                    plotter.plot(data[data_set]['data'], color=data[data_set]["properties"]["color"], label=data[data_set]["properties"]["label"])
                plotter.xlabel(xlabel=plot_properties["xlabel"], weight=plot_properties["label_weight"], style=plot_properties["label_style"])
                plotter.ylabel(ylabel=plot_properties["ylabel"], weight=plot_properties["label_weight"], style=plot_properties["label_style"])
                plotter.title(label=plot_properties["title"], weight=plot_properties["title_weight"], style=plot_properties["title_style"])
                plotter.grid(visible=plot_properties["grid"])
                plotter.legend()

                fig = plotter.gcf()
                fig.set_size_inches(h = int(plot_properties["height"])/2.54, w = int(plot_properties["width"])/2.54, forward = True)
                plotter.show()

                ans = input_catch("Are you ok with this plot? (y/n)?", "n")
                if ans.lower() in ("y", "yes", "yep"):
                    fig.savefig("output.png", dpi=int(plot_properties["dpi"]))
                plotter.clf()
                plotter.close(None)
                del fig
            else:
                print("Use \"config plot\" command to configure plotter properties..")
        else:
            print_help()
