import matplotlib.pyplot as plotter
from openpyxl import load_workbook
import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter.colorchooser import askcolor
from statistics import mean, stdev 

def print_help():
    print("Plot generator from excel options (2023) Mateusz Ferenc:")
    print("\tq, quit, exit, leave - to exit")
    print("\tload - open file dialog window to select file to load data from")
    print("\tunload - delete loaded file from temporary memory")
    print("\tselect <row-start> <row-end> <column-start> <column-end> - select data within given range (column accepts excel column representation)")
    print("\tdata or data <data-set> - print loaded datasets (only informations) / print contents of <data-set>")
    print("\tclear or clear <data-set> - clear whole database / clear <data-set> entry")
    print("\tconfig plot or config data or config <data-set> - configure plotter properties / configure each dataset plot properties / configure <data-set> plot properties")
    print("\tgenerate - generate plot from given data")

def excel_column_to_num(col: str) -> (None | int):
    if type(col) is not str:
        return None
    col_num = 0
    for c in range(len(col)):
        ascii_ = ord(col[-(1 + c)])
        if 65 <= ascii_ <= 90:
            col_num += (ascii_ - 65) * pow(26, c)
        else:
            return None
    return col_num

def input_catch(text: str):
    try:
        temp = input(text)
        return temp
    except KeyboardInterrupt:
        return None

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()

    workbook = None
    current_sheet = None
    data = {}
    plot_properties = {
        "dpi": 500,
        "xlabel": None,
        "ylabel": None,
        "label_weight": "light",
        "label_style": "italic",
        "title": None,
        "title_weight": "bold",
        "title_style": "bold",
        "grid": False
    }
    plot_properties_ok = False

    while(True):
        command = input(">")
        command, command_data = command.split()[0].lower(), command.split()[1:]
        if command in ("h", "help"):
            print_help()
        elif command in ("q", "quit", "exit", "leave"):
            if workbook is not None:
                workbook.close
            exit()
        elif command == "load":
            ans = ""
            if workbook is not None:
                ans = input(f"{current_sheet.title} is currently loaded, unload? (y/n)")
            if ans.lower() in ("y", "yes", "yep") or workbook is None:
                path = askopenfilename(title="Choose excel file", filetypes=[("Excel files", ".xlsx")], parent=root)
                try:
                    workbook = load_workbook(path)
                except FileNotFoundError:
                    print(f"{path} file not found...")
                except Exception as e:
                    print(f"{e}")                   
                else:
                    current_sheet = workbook.active
                    print(f"{current_sheet.title} loaded")
        elif command == "unload":
            if workbook is not None:
                workbook.remove()
                print(f"{current_sheet.title} unloaded")
                workbook = None
            current_sheet = None
        elif command == "select":
            if len(command_data) == 4:
                row_from, row_to, col_from, col_to = int(command_data[0]), int(command_data[1]), command_data[2], command_data[3]
                col_ok = True
                temp = excel_column_to_num(col_from)
                col_from = temp + 1
                if temp is None:
                    col_ok = False
                temp = excel_column_to_num(col_to)
                col_from = temp + 1
                if temp is None:
                    col_ok =  False
                temp_data = []
                if row_from <= row_to and col_from <= col_to:
                    try:
                        for row in range(row_from, row_to + 1):
                            for col in range(col_from, col_to + 1):
                                value = current_sheet.cell(row=row, column=col).value
                                if type(value) is int or type(value) is float:
                                    temp_data.append(round(value, 4))
                                else:
                                    print(f"value: {value} in row: {row}, col: {col} is not int of float type\nType: {type(value)}\nRemoving selected data...")
                                    raise Exception("data is not valid")
                    except:
                        pass
                    else:
                        print(f"Data is valid.\nLoaded datasets: {datasets}")
                        datasets = [d for d in data.keys()]
                        can_exit = False
                        while not can_exit:
                            data_name = input("Enter name of new data set?")
                            if data_name not in ('plot', 'data'):
                                data[data_name] = {}
                                can_exit = True
                            else:
                                print(f"{data_name} is protected name, use another..")
                        data[data_name]["data"] = temp_data
                        data[data_name]["info"] = {}
                        data[data_name]["info"]["sheet_name"] = current_sheet.title
                        data[data_name]["info"]["row_from_to"] = f"{row_from} : {row_to}"
                        data[data_name]["info"]["col_from_to"] = f"{command_data[2]} : {command_data[3]}"
                        label = input("Enter data label\n?")
                        data[data_name]["properites"] = {}
                        data[data_name]["properites"]["label"] = label
                        print("Color pick window summoned..")
                        color = askcolor(title="Choose plot color", parent=root)
                        data[data_name]["properites"]["color"] = color[1] if color != (None, None) else "#000000"
                else:
                    print("Selected range is not valid...\nTry again")
            else:
                print("Use: select <row-start> <row-end> <column-start> <column-end>")
        elif command == "data":
            if len(command_data) == 0:
                if len(data):
                    for data_set in data.keys():
                        print(f"Dataset: \"{data_set}\" of size: {len(data[data_set]['data'])} elements")
                        print(f"Sheet name: {data[data_set]['info']['sheet_name']}, data selected= rows range {data[data_set]['info']['row_from_to']}, columns range {data[data_set]['info']['col_from_to']}")
                        print(f"Properties - label: {data[data_set]['properties']['label']}, color: {data[data_set]['properties']['color']}")
                        temp = data[data_set]['data']
                        print(f"Statistics - min: {min(temp)}, max: {max(temp)}, avg: {mean(temp)}, stdev: {stdev(temp)}")
                else:
                    print(f"No data loaded.\nUse load and select commands to load new data")
        elif command == "clear":
            if len(command_data) == 0:
                data = {}
                plot_properties_ok = False
                print("All data cleared")
            elif len(command_data) == 1:
                if command_data[0] in data.keys():
                    del data[command_data[0]]
                    print(f"\"{command_data[0]}\" entries was removed from database")
            else:
                print("Use: clear or clear <data-set>")
        elif command == "config":
            if len(command_data) == 1:
                if command_data[0] == "plot":
                    can_exit = False
                    while not can_exit:
                        plot_properties["dpi"] = int(input("Enter plot DPI\n?"))
                        plot_properties["xlabel"] = input("Enter x (horizontal) label name\n?")
                        plot_properties["ylabel"] = input("Enter y (vertical) label name\n?")
                        font_weights = ('ultralight', 'light', 'normal', 'regular', 'book', 'medium', 'roman', 'semibold', 'demibold', 'demi', 'bold', 'heavy', 'extra bold', 'black')
                        w = input(f"Enter labels font weight (valid: {font_weights})\n?")
                        if w in font_weights:
                            plot_properties["label_weight"] = w
                        else:
                            print(f"fontweight: {w} is not valid..\n\"bold\" used as default")
                            plot_properties["label_weight"] = "bold"
                        font_styles = ('normal', 'italic', 'oblique')
                        s = input(f"Enter labels font style (valid: {font_styles})\n?")
                        if s in font_styles:
                            plot_properties["label_style"] = s
                        else:
                            print(f"fontstyle: {s} is not valid..\n\"light\" used as default")
                            plot_properties["label_style"] = "light"
                        plot_properties["title"] = input("Enter plot title\n?")
                        w = input(f"Enter title font weight (valid: {font_weights})\n?")
                        if w in font_weights:
                            plot_properties["title_weight"] = w
                        else:
                            print(f"fontweight: {w} is not valid..\n\"bold\" used as default")
                            plot_properties["title_weight"] = "bold"
                        s = input(f"Enter labels font style (valid: {font_styles})\n?")
                        if s in font_styles:
                            plot_properties["title_style"] = s
                        else:
                            print(f"fontstyle: {s} is not valid..\n\"blod\" used as default")
                            plot_properties["title_style"] = "bold"
                        g = input("Enable grid? (valid: True, False)\n?")
                        if g in ("True", "False"):
                            plot_properties["grid"] = g
                        else:
                            print(f"option: {g} is not valid..\n\"False\" used as default")
                        print()
                        for key in plot_properties.keys():
                            print(f"{key}: {plot_properties[key]}")
                        ans = input("Are you ok with these values? (y/n)?")
                        if ans.lower() in ("y", "yes", "yep"):
                            can_exit = True
                            plot_properties_ok = True
                            print("OK!")
                elif command_data[0] == "data":
                    can_exit = False
                    while not can_exit:
                        for ds_name in data.keys():
                            print(f"{ds_name}:")
                            label = data[ds_name[0]]["properites"]["label"]
                            from_input = input_catch(f"Enter data label for \"{ds_name}\" or preserve old {label} using ctrl+c\n?")
                            data[ds_name]["properites"]["label"] = from_input if from_input is not None else label
                            color = data[ds_name[0]]["properites"]["color"]
                            selected_color = askcolor(title=f"Choose plot color, currently: {color}", parent=root)
                            data[ds_name]["properites"]["color"] = selected_color[1] if selected_color != (None, None) else color
                        for key in data.keys():
                            print(f"{key}:")
                            for k in data[key]["properites"].keys():
                                print(f"\t{k}: {data[key]['properites'][k]}")
                        ans = input("Are you ok with these values? (y/n)?")
                        if ans.lower() in ("y", "yes", "yep"):
                            can_exit = True
                            print("OK!")
                elif command_data[0] in data.keys():
                    can_exit = False
                    while not can_exit:
                        print(f"{command_data[0]}:")
                        label = data[command_data[0]]["properites"]["label"]
                        from_input = input_catch(f"Enter data label for \"{ds_name}\" or preserve old {label} using ctrl+c\n?")
                        data[data_name]["properites"]["label"] = from_input if from_input is not None else label
                        color = data[command_data[0]]["properites"]["color"]
                        selected_color = askcolor(title=f"Choose plot color, currently: {color}", parent=root)
                        data[data_name]["properites"]["color"] = selected_color[1] if selected_color != (None, None) else color
            else:
                print("Use: config plot or config data or config <data-set>")
        elif command == "generate":
            if plot_properties_ok and len(data):
                for data_set in data.keys():
                    plotter.plot(data[data_set]['data'], color=data[data_set]["properites"]["color"], label=data[data_set]["properites"]["label"])
                plotter.xlabel(xlabel=plot_properties["xlabel"], weight=plot_properties["label_weight"], style=plot_properties["label_style"])
                plotter.ylabel(ylabel=plot_properties["ylabel"], weight=plot_properties["label_weight"], style=plot_properties["label_style"])
                plotter.title(label=plot_properties["title"], weight=plot_properties["title_weight"], style=plot_properties["title_style"])
                plotter.grid(visible=plot_properties["grid"])
                plotter.legend()

                fig = plotter.gcf()
                plotter.show()

                ans = input("Are you ok with this plot? (y/n)?")
                if ans.lower() in ("y", "yes", "yep"):
                    fig.savefig("output.png", dpi=int(plot_properties["dpi"]))
                plotter.clf()
                plotter.close(None)
                del fig
            else:
                print("Use \"config plot\" command to configure plotter properites..")
        elif command.startswith("exec"): # debug-only
            try:
                exec(command.replace("exec ", ""))
            except Exception as e:
                print(e)
